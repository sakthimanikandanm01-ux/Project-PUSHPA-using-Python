import cv2
from ultralytics import YOLO
import cvzone
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import requests
import pandas as pd
import os
import time
from twilio.rest import Client
import csv
from openpyxl import load_workbook
from email.mime.image import MIMEImage
import ssl
from urllib3.poolmanager import PoolManager
from requests.adapters import HTTPAdapter
import threading



# === Configuration ===
model = YOLO('best.pt')
names = model.names  # ['tiger', 'elephant', 'man', 'weapon']

SENDER_EMAIL = "rajinisakthi48@gmail.com"
RECEIVER_EMAILS = ["sakthimanikandanm01@gmail.com","akilraj833@gmail.com"]
EMAIL_PASSWORD = "dupjnhrqbrtkcwts"



# === Global flags ===
EMAIL_SENT = False
TIGER_ALERT_SENT = False
ELEPHANT_ALERT_SENT = False
tiger_timer_start = None
elephant_timer_start = None

# === Utility Functions ===
def send_email_alert(counts, species=None, frame=None):
    subject = f"[Wildlife Alert] {species.upper() if species else 'Unknown'} Detected"
    body = f"📢 Wildlife Detection Alert\nTime: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\nDetected:\n"
    for key, val in counts.items():
        if val > 0:
            body += f"- {key.capitalize()}: {val}\n"

    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(RECEIVER_EMAILS)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    # === Save and attach screenshot ===
    if frame is not None:
        image_path = "detected_frame.jpg"
        cv2.imwrite(image_path, frame)

        with open(image_path, 'rb') as f:
            img_data = f.read()
            image = MIMEImage(img_data, name=os.path.basename(image_path))
            msg.attach(image)

    # === Send email ===
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, EMAIL_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAILS, msg.as_string())
        server.quit()
        print(f"✅ Email alert with image sent for {species}")
    except Exception as e:
        print("❌ Email failed:", str(e))


def send_email_alert_threaded(counts, species=None, frame=None):
    thread = threading.Thread(target=send_email_alert, args=(counts, species, frame))
    thread.start()





from twilio.rest import Client

def send_whatsapp_alert(counts, species=None, frame=None):
    from twilio.rest import Client

    account_sid = 'ACd080076022eb4ed3472b46e938aad325'
    auth_token = '0c4bc1e0cb78d93917b07ab5e257db81'
    client = Client(account_sid, auth_token)

    recipient = 'whatsapp:+919787494341'
    twilio_number = 'whatsapp:+14155238886'

    time_str = time.strftime('%Y-%m-%d %H:%M:%S')
    detected_text = "\n".join([f"- {k.capitalize()}: {v}" for k, v in counts.items() if v > 0])

    if not detected_text:
        print("🟡 No detected object to alert via WhatsApp.")
        return

    image_path = "detected_frame.jpg"
    if frame is not None:
        cv2.imwrite(image_path, frame)
        image_url = upload_image_to_imgur(image_path)
    else:
        image_url = None

    message_text = f"""
📢 Wildlife Detection Alert
Time: {time_str}
Species: {species.capitalize() if species else 'Unknown'}

Detected:
{detected_text}
""".strip()

    try:
        message = client.messages.create(
            body=message_text,
            from_=twilio_number,
            to=recipient,
            media_url=[image_url] if image_url else None
        )
        print("✅ WhatsApp message sent with image! SID:", message.sid)
    except Exception as e:
        print("❌ Failed to send WhatsApp alert:", str(e))

def send_whatsapp_alert_threaded(counts, species=None, frame=None):
    thread = threading.Thread(target=send_whatsapp_alert, args=(counts, species, frame))
    thread.start()


class SSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = ssl.create_default_context()
        context.set_ciphers("DEFAULT@SECLEVEL=1")  # ↓ Security level
        kwargs['ssl_context'] = context
        return super().init_poolmanager(*args, **kwargs)

def upload_image_to_imgur(image_path):
    IMGUR_CLIENT_ID = "546f6c2b2c9fd1d"  # Replace with your own client ID

    session = requests.Session()
    session.mount("https://", SSLAdapter())

    headers = {"Authorization": f"Client-ID {IMGUR_CLIENT_ID}"}
    with open(image_path, 'rb') as f:
        image_data = {'image': f}
        try:
            response = session.post(
                url="https://api.imgur.com/3/upload",
                headers=headers,
                files=image_data,
                timeout=10
            )
            if response.status_code == 200:
                return response.json()['data']['link']
            else:
                print("❌ Imgur upload failed:", response.text)
                return None
        except requests.exceptions.SSLError as ssl_err:
            print("❌ SSL Error during Imgur upload:", ssl_err)
            return None
        except Exception as e:
            print("❌ General error during Imgur upload:", e)
            return None




def log_to_temp_xlsx(counts, species, filename="wildlife_log_temp.xlsx"):
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    row = {"Time": timestamp, "Species": species}
    row.update({k: v for k, v in counts.items() if v > 0})

    df = pd.DataFrame([row])

    if not os.path.exists(filename):
        # Create new file with header
        df.to_excel(filename, index=False, engine='openpyxl')
    else:
        # Append to existing file
        wb = load_workbook(filename)
        ws = wb.active
        startrow = ws.max_row

        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=startrow)

    print(f"✅ Logged to Excel (real .xlsx): {filename}")


def log_to_temp_xlsx_threaded(counts, species, filename="wildlife_log_temp.xlsx"):
    thread = threading.Thread(target=log_to_temp_xlsx, args=(counts, species, filename))
    thread.start()

# === Detection Loop ===
cap = cv2.VideoCapture("fireu.mp4")  # or use 0 for webcam
frame_count = 0

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame_count += 1
    if frame_count % 3 != 0:
        continue

    frame = cv2.resize(frame, (1020, 500))
    results = model.track(frame, persist=True, classes=[0, 1, 2, 3])
    count_dict = {'tiger': 0, 'elephant': 0, 'man': 0, 'weapon': 0}

    if results[0].boxes.id is not None:
        ids = results[0].boxes.id.cpu().numpy().astype(int)
        boxes = results[0].boxes.xyxy.cpu().numpy().astype(int)
        class_ids = results[0].boxes.cls.int().cpu().tolist()
        current_time = time.time()

        for track_id, box, class_id in zip(ids, boxes, class_ids):
            x1, y1, x2, y2 = box
            label = names[class_id]
            if label in count_dict:
                count_dict[label] += 1

            if label == 'tiger':
                if tiger_timer_start is None:
                    tiger_timer_start = current_time
                elif not TIGER_ALERT_SENT and (current_time - tiger_timer_start >= 3):
                    image_path = "detected_frame.jpg"
                    cv2.imwrite(image_path, frame)

                    send_email_alert_threaded(count_dict, species="tiger",frame=frame)
                    send_whatsapp_alert_threaded(count_dict, species="tiger",frame=frame)
                    log_to_temp_xlsx_threaded(count_dict, species="tiger")
                    TIGER_ALERT_SENT = True
            else:
                tiger_timer_start = None
                TIGER_ALERT_SENT = False

            if label == 'elephant':
                if elephant_timer_start is None:
                    elephant_timer_start = current_time
                elif not ELEPHANT_ALERT_SENT and (current_time - elephant_timer_start >= 3):
                    image_path = "detected_frame.jpg"
                    cv2.imwrite(image_path, frame)

                    send_email_alert_threaded(count_dict, species="elephant",frame=frame)
                    send_whatsapp_alert_threaded(count_dict, species="elephant",frame=frame)
                    log_to_temp_xlsx_threaded(count_dict, species="elephant")
                    ELEPHANT_ALERT_SENT = True
            else:
                elephant_timer_start = None
                ELEPHANT_ALERT_SENT = False

            if label == 'weapon' and not EMAIL_SENT:
                image_path = "detected_frame.jpg"
                cv2.imwrite(image_path, frame)

                send_email_alert_threaded(count_dict, species="weapon",frame=frame)
                send_whatsapp_alert_threaded(count_dict, species="weapon",frame=frame)
                log_to_temp_xlsx_threaded(count_dict, species="weapon")
                EMAIL_SENT = True

            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(frame, f'{label} ID:{track_id}', (x1 + 5, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

    y_offset = 20
    for key, value in count_dict.items():
        cv2.putText(frame, f'{key}: {value}', (10, y_offset),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        y_offset += 25

    cv2.imshow("RGB", frame)
    if cv2.waitKey(1) & 0xFF == 27:
        break

cap.release()
cv2.destroyAllWindows()

#.\venv\Scripts\activate
#python main.py
#cd C:\Users\hellb\Downloads\pushpa
#written by
#Sakthi Manikandan T
#dupjnhrqbrtkcwts-app pass
#account_sid = 'ACd080076022eb4ed3472b46e938aad325'
#auth_token = '0c4bc1e0cb78d93917b07ab5e257db81'
